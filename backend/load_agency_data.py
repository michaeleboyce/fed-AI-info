#!/usr/bin/env python3
"""
Load Federal Agency GenAI Provisioning data from Excel into SQLite database.

This script processes the Excel file containing information about how federal agencies
are adopting AI internally (staff LLMs, coding assistants, specialized tools).
"""

import openpyxl
import sqlite3
from datetime import datetime
from pathlib import Path

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / 'data'
DB_PATH = DATA_DIR / 'fedramp.db'
EXCEL_PATH = Path.home() / 'Downloads' / 'Federal_Agency_GenAI_Provisioning.xlsx'

def create_tables(conn):
    """Create database tables for agency AI usage data."""
    cursor = conn.cursor()

    # Main agency AI usage table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS agency_ai_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            agency_name TEXT NOT NULL,
            agency_category TEXT NOT NULL,  -- 'staff_llm' or 'specialized'
            has_staff_llm TEXT,
            llm_name TEXT,
            has_coding_assistant TEXT,
            scope TEXT,
            solution_type TEXT,
            non_public_allowed TEXT,
            other_ai_present TEXT,
            tool_name TEXT,  -- for specialized AI
            tool_purpose TEXT,
            notes TEXT,
            sources TEXT,
            analyzed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            slug TEXT  -- URL-friendly agency identifier
        )
    ''')

    conn.commit()
    print("✅ Database tables created")

def generate_slug(agency_name):
    """Generate URL-friendly slug from agency name."""
    import re
    # Convert to lowercase, replace spaces and special chars with hyphens
    slug = agency_name.lower()
    slug = re.sub(r'[^\w\s-]', '', slug)
    slug = re.sub(r'[-\s]+', '-', slug)
    return slug.strip('-')

def load_staff_llm_data(conn, wb):
    """Load staff LLM and coding assistant data."""
    cursor = conn.cursor()
    ws = wb['Staff LLMs & Coding']

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue

        data = dict(zip(headers, row))

        # Extract LLM name from notes if mentioned
        llm_name = None
        notes = data.get('Notes/Comments', '')
        if notes and ('GPT' in notes or 'Chat' in notes):
            # Try to extract tool name from quotes
            import re
            match = re.search(r"['\u2018\u2019]([\w\s-]+GPT|[\w\s-]+Chat)[\u2019']", notes)
            if match:
                llm_name = match.group(1)

        cursor.execute('''
            INSERT INTO agency_ai_usage
            (agency_name, agency_category, has_staff_llm, llm_name, has_coding_assistant,
             scope, solution_type, non_public_allowed, other_ai_present, notes, sources, slug)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('Agency/Department'),
            'staff_llm',
            data.get('Has staff LLM chatbot?'),
            llm_name,
            data.get('Has AI coding assistant?'),
            data.get('Scope'),
            data.get('Solution type'),
            data.get('Non-public info allowed?'),
            data.get('Other AI (non‑chat) present?'),
            data.get('Notes/Comments'),
            data.get('Sources'),
            generate_slug(data.get('Agency/Department', ''))
        ))
        count += 1

    conn.commit()
    print(f"✅ Loaded {count} staff LLM/coding assistant entries")

def load_specialized_ai_data(conn, wb):
    """Load specialized AI (non-chat) data."""
    cursor = conn.cursor()
    ws = wb['Specialized AI (Non-chat)']

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue

        data = dict(zip(headers, row))

        cursor.execute('''
            INSERT INTO agency_ai_usage
            (agency_name, agency_category, tool_name, tool_purpose,
             solution_type, scope, non_public_allowed, sources, slug)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('Agency/Department'),
            'specialized',
            data.get('Tool / Capability'),
            data.get('Purpose'),
            data.get('Custom or Commercial'),
            data.get('Scope'),
            data.get('Non-public info allowed?'),
            data.get('Sources'),
            generate_slug(data.get('Agency/Department', ''))
        ))
        count += 1

    conn.commit()
    print(f"✅ Loaded {count} specialized AI tool entries")

def main():
    """Main execution function."""
    print("🚀 Loading Federal Agency AI data into database...")
    print(f"   Excel file: {EXCEL_PATH}")
    print(f"   Database: {DB_PATH}")
    print()

    # Check if Excel file exists
    if not EXCEL_PATH.exists():
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        print("   Please ensure the file is in your Downloads folder")
        return 1

    # Open Excel file
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"📊 Excel sheets found: {wb.sheetnames}")
    print()

    # Connect to database
    conn = sqlite3.connect(DB_PATH)

    # Create tables
    create_tables(conn)
    print()

    # Clear existing data (optional - comment out to keep existing data)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM agency_ai_usage')
    conn.commit()
    print("🗑️  Cleared existing agency data")
    print()

    # Load data
    load_staff_llm_data(conn, wb)
    load_specialized_ai_data(conn, wb)
    print()

    # Show summary
    cursor.execute('SELECT COUNT(*) FROM agency_ai_usage')
    total = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM agency_ai_usage WHERE agency_category = 'staff_llm'")
    staff_llm_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM agency_ai_usage WHERE agency_category = 'specialized'")
    specialized_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT agency_name) FROM agency_ai_usage")
    unique_agencies = cursor.fetchone()[0]

    print("📈 Summary:")
    print(f"   Total entries: {total}")
    print(f"   Staff LLM/Coding entries: {staff_llm_count}")
    print(f"   Specialized AI entries: {specialized_count}")
    print(f"   Unique agencies: {unique_agencies}")
    print()

    # Show sample
    print("📋 Sample entries:")
    cursor.execute('''
        SELECT agency_name, has_staff_llm, llm_name, has_coding_assistant, solution_type
        FROM agency_ai_usage
        WHERE agency_category = 'staff_llm'
        LIMIT 5
    ''')
    for row in cursor.fetchall():
        print(f"   • {row[0]}: LLM={row[1]} ({row[2]}), Coding={row[3]}, Type={row[4]}")

    conn.close()
    print()
    print("✅ Done! Agency AI data loaded successfully")
    return 0

if __name__ == '__main__':
    exit(main())
